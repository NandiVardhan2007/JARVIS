"""
VISION Dynamic Excel Task Tracker Engine.
Generates a colorful, formula-driven Excel Workbook with embedded Pie/Donut charts,
KPI dashboard cards, 12 monthly sheets with 31-day tracking grids, and conditional formatting.
"""

import os
from datetime import datetime
from typing import Optional, Dict, Any, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference
from openpyxl.formatting.rule import FormulaRule, CellIsRule

from vision.memory.task_tracker_db import task_db
from vision.logger import logger

EXCEL_OUTPUT_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "data", "VISION_Task_Tracker.xlsx")

MONTHS = ["January", "February", "March", "April", "May", "June", 
          "July", "August", "September", "October", "November", "December"]
MONTH_SHORT = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


class ExcelTrackerEngine:
    def __init__(self, filepath: str = EXCEL_OUTPUT_PATH):
        self.filepath = filepath
        os.makedirs(os.path.dirname(self.filepath), exist_ok=True)

    def generate_workbook(self, year: Optional[int] = None) -> str:
        """Build and save the entire interactive Excel Task Tracker workbook."""
        curr = task_db.get_current_date_info()
        year = year or curr["year"]
        task_db.ensure_daily_leetcode_tasks(year=year)

        wb = openpyxl.Workbook()
        # Default sheet will be our Dashboard
        ws_dash = wb.active
        ws_dash.title = "VISION_DASHBOARD"
        ws_dash.views.sheetView[0].showGridLines = True

        # Styles definition
        font_main = Font(name="Segoe UI", size=11, color="FFFFFF")
        font_title = Font(name="Segoe UI", size=18, bold=True, color="00F2FE")
        font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="A0AEC0")
        font_card_num = Font(name="Segoe UI", size=22, bold=True, color="FFFFFF")
        font_card_label = Font(name="Segoe UI", size=9, bold=True, color="CBD5E0")
        font_tbl_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")

        fill_dark_bg = PatternFill(start_color="0A0F1D", end_color="0A0F1D", fill_type="solid")
        fill_card_cyan = PatternFill(start_color="005274", end_color="005274", fill_type="solid")
        fill_card_green = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")
        fill_card_purple = PatternFill(start_color="4C1D95", end_color="4C1D95", fill_type="solid")
        fill_card_amber = PatternFill(start_color="78350F", end_color="78350F", fill_type="solid")
        fill_card_blue = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")

        fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        fill_row_even = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        fill_row_odd = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

        thin_border_side = Side(border_style="thin", color="334155")
        grid_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        # ── 1. Render Dashboard Sheet ──
        # Header Banner
        ws_dash.merge_cells("B2:K2")
        ws_dash["B2"] = f"⚡ VISION AUTONOMOUS TASK TRACKER & PRODUCTIVITY OS — {year}"
        ws_dash["B2"].font = font_title
        ws_dash["B2"].alignment = align_left

        lc_streak = task_db.calculate_leetcode_streak(year)
        ws_dash.merge_cells("B3:K3")
        ws_dash["B3"] = f"LeetCode & Coding Habit: Active | Streak: {lc_streak} Days 🔥 | Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws_dash["B3"].font = font_subtitle
        ws_dash["B3"].alignment = align_left

        # Top KPI Summary Cards (B5:C6, D5:E6, F5:G6, H5:I6, J5:K6)
        # Note: B10:H21 is the monthly table. C is Total, D is Completed, E is Pending.
        cards = [
            ("TOTAL TASKS", "=SUM(C10:C21)", fill_card_blue, "B", "C"),
            ("COMPLETED", "=SUM(D10:D21)", fill_card_green, "D", "E"),
            ("PENDING", "=SUM(E10:E21)", fill_card_amber, "F", "G"),
            ("COMPLETION RATE", '=IF(B5>0, D5/B5, 0)', fill_card_cyan, "H", "I"),
            ("CURRENT STREAK", f"{task_db.calculate_streak(year)} Days 🔥", fill_card_purple, "J", "K")
        ]

        for label, val_formula, fill_c, col1, col2 in cards:
            ws_dash.merge_cells(f"{col1}5:{col2}5")
            ws_dash[f"{col1}5"] = val_formula
            ws_dash[f"{col1}5"].font = font_card_num
            ws_dash[f"{col1}5"].fill = fill_c
            ws_dash[f"{col1}5"].alignment = align_center
            if label == "COMPLETION RATE":
                ws_dash[f"{col1}5"].number_format = '0.0%'

            ws_dash.merge_cells(f"{col1}6:{col2}6")
            ws_dash[f"{col1}6"] = label
            ws_dash[f"{col1}6"].font = font_card_label
            ws_dash[f"{col1}6"].fill = fill_c
            ws_dash[f"{col1}6"].alignment = align_center

            for r in range(5, 7):
                for col in [col1, col2]:
                    ws_dash[f"{col}{r}"].border = grid_border

        # Monthly Performance Table (B9:H22)
        ws_dash.merge_cells("B8:H8")
        ws_dash["B8"] = "📅 MONTHLY PROGRESS BREAKDOWN"
        ws_dash["B8"].font = Font(name="Segoe UI", size=12, bold=True, color="00F2FE")
        ws_dash["B8"].alignment = align_left

        headers = ["Month", "Total Tasks", "Completed", "Pending", "Success Rate", "Status Gauge", "Active"]
        for col_idx, h in enumerate(headers, start=2):
            cell = ws_dash.cell(row=9, column=col_idx, value=h)
            cell.font = font_tbl_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = grid_border

        for m_idx, (m_full, m_sh) in enumerate(zip(MONTHS, MONTH_SHORT), start=1):
            row_idx = 9 + m_idx
            is_active = "⚡ CURRENT" if m_full == curr["month"] else ""
            
            row_fill = fill_row_odd if m_idx % 2 == 1 else fill_row_even

            # Month Name
            c1 = ws_dash.cell(row=row_idx, column=2, value=m_full)
            # Total
            c2 = ws_dash.cell(row=row_idx, column=3, value=f"='{m_sh}'!G3")
            # Completed
            c3 = ws_dash.cell(row=row_idx, column=4, value=f"='{m_sh}'!G4")
            # Pending
            c4 = ws_dash.cell(row=row_idx, column=5, value=f"=C{row_idx}-D{row_idx}")
            # Success Rate
            c5 = ws_dash.cell(row=row_idx, column=6, value=f'=IF(C{row_idx}>0, D{row_idx}/C{row_idx}, 0)')
            c5.number_format = '0.0%'
            # Status Visual
            c6 = ws_dash.cell(row=row_idx, column=7, value=f'=IF(C{row_idx}=0, "⚪ NO TASKS", IF(D{row_idx}=C{row_idx}, "🌟 PERFECT", IF(F{row_idx}>=0.7, "🟢 ON TRACK", "🟡 IN PROGRESS")))')
            # Active Indicator
            c7 = ws_dash.cell(row=row_idx, column=8, value=is_active)

            for col_i in range(2, 9):
                cell = ws_dash.cell(row=row_idx, column=col_i)
                cell.font = Font(name="Segoe UI", size=10, color="FFFFFF", bold=(col_i in [2, 6, 8]))
                cell.fill = row_fill
                cell.border = grid_border
                if col_i == 2:
                    cell.alignment = align_left
                else:
                    cell.alignment = align_center

        # Pie Chart Data Source Table (J9:K11)
        ws_dash.cell(row=9, column=10, value="Status").font = font_tbl_header
        ws_dash.cell(row=9, column=10).fill = fill_header
        ws_dash.cell(row=9, column=11, value="Count").font = font_tbl_header
        ws_dash.cell(row=9, column=11).fill = fill_header

        ws_dash.cell(row=10, column=10, value="Completed").font = Font(color="00FF88", bold=True)
        ws_dash.cell(row=10, column=11, value="=D5").font = Font(color="FFFFFF")

        ws_dash.cell(row=11, column=10, value="Pending").font = Font(color="FFAA00", bold=True)
        ws_dash.cell(row=11, column=11, value="=F5").font = Font(color="FFFFFF")

        for r in range(9, 12):
            for c in range(10, 12):
                ws_dash.cell(row=r, column=c).border = grid_border

        # Embed OpenPyXL Pie Chart
        try:
            pie = PieChart()
            pie.title = "🏆 Overall Task Completion Ratio"
            pie.title.text.font = font_tbl_header
            labels = Reference(ws_dash, min_col=10, min_row=10, max_row=11)
            data = Reference(ws_dash, min_col=11, min_row=9, max_row=11)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.width = 14
            pie.height = 8.5
            pie.style = 10
            ws_dash.add_chart(pie, "J13")
        except Exception as e:
            logger.warning(f"[ExcelTracker] Pie chart creation warning: {e}")

        # Set Dashboard column widths
        col_widths = {
            "A": 3, "B": 16, "C": 14, "D": 14, "E": 14, "F": 16, "G": 18, "H": 16, "I": 3, "J": 15, "K": 15
        }
        for col_l, width in col_widths.items():
            ws_dash.column_dimensions[col_l].width = width

        # ── 2. Render 12 Monthly Tracking Sheets ──
        tasks_all = task_db.get_all_tasks(year=year)
        tasks_by_month: Dict[int, List[Dict[str, Any]]] = {m: [] for m in range(1, 13)}
        for t in tasks_all:
            m_num = t["month_num"]
            if m_num in tasks_by_month:
                tasks_by_month[m_num].append(t)

        for m_num, (m_name, m_short) in enumerate(zip(MONTHS, MONTH_SHORT), start=1):
            ws_month = wb.create_sheet(title=m_short)
            ws_month.views.sheetView[0].showGridLines = True

            # Sheet Title
            ws_month.merge_cells("A1:E1")
            ws_month["A1"] = f"🗓️ {m_name.upper()} {year} — DAILY TASK TRACKER (DAYS 1–31)"
            ws_month["A1"].font = Font(name="Segoe UI", size=14, bold=True, color="00F2FE")
            ws_month["A1"].alignment = align_left

            # Quick Metric Labels for Dashboard Links
            ws_month["F3"] = "Total Tasks:"
            ws_month["G3"] = "=COUNTA(B8:B150)"
            ws_month["F4"] = "Completed:"
            ws_month["G4"] = '=COUNTIF(E8:E150, TRUE) + COUNTIF(E8:E150, "TRUE") + COUNTIF(E8:E150, 1)'
            
            for cell_ref in ["F3", "F4"]:
                ws_month[cell_ref].font = Font(name="Segoe UI", size=9, bold=True, color="94A3B8")
            for cell_ref in ["G3", "G4"]:
                ws_month[cell_ref].font = Font(name="Segoe UI", size=10, bold=True, color="00FF88")

            # Table Header
            month_headers = ["Day (1-31)", "Task Title / Objective", "Category", "Priority", "Completed (TRUE/FALSE)", "Time Logged"]
            for col_i, h in enumerate(month_headers, start=1):
                cell = ws_month.cell(row=7, column=col_i, value=h)
                cell.font = font_tbl_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = grid_border

            m_tasks = tasks_by_month.get(m_num, [])
            
            # If no tasks exist for this month yet, provide clean pre-populated sample rows
            if not m_tasks and m_name == curr["month"]:
                # Seed current month with initial sample tasks
                sample_tasks = [
                    {"day": curr["day"], "title": "Morning Routine & Goal Alignment", "category": "Habits", "priority": "High", "is_completed": 1, "created_at": datetime.now().strftime("%H:%M")},
                    {"day": curr["day"], "title": "Review VISION Architecture & Automation", "category": "Coding", "priority": "High", "is_completed": 0, "created_at": datetime.now().strftime("%H:%M")},
                    {"day": curr["day"], "title": "Evening Workout & Hydration", "category": "Fitness", "priority": "Medium", "is_completed": 0, "created_at": datetime.now().strftime("%H:%M")}
                ]
                for st in sample_tasks:
                    task_db.add_task(
                        title=st["title"], 
                        day=st["day"], 
                        month=m_name, 
                        year=year, 
                        category=st["category"], 
                        priority=st["priority"]
                    )
                m_tasks = task_db.get_tasks_for_month(month=m_name, year=year)

            row_ptr = 8
            for t in m_tasks:
                r_fill = fill_row_odd if (row_ptr % 2 == 1) else fill_row_even

                c_day = ws_month.cell(row=row_ptr, column=1, value=f"Day {t['day']:02d}")
                c_title = ws_month.cell(row=row_ptr, column=2, value=t["title"])
                c_cat = ws_month.cell(row=row_ptr, column=3, value=t["category"] or "General")
                c_prio = ws_month.cell(row=row_ptr, column=4, value=t["priority"] or "Medium")
                c_done = ws_month.cell(row=row_ptr, column=5, value=True if t["is_completed"] == 1 else False)
                raw_time = t.get("completed_at") or t.get("created_at") or ""
                c_time = ws_month.cell(row=row_ptr, column=6, value=raw_time[:16])

                for col_i in range(1, 7):
                    cell = ws_month.cell(row=row_ptr, column=col_i)
                    cell.font = Font(name="Segoe UI", size=10, color="FFFFFF")
                    cell.fill = r_fill
                    cell.border = grid_border
                    if col_i in [1, 3, 4, 5, 6]:
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_left

                row_ptr += 1

            # Fill up to at least 25 clean rows for smooth manual Excel entry
            while row_ptr < 32:
                r_fill = fill_row_odd if (row_ptr % 2 == 1) else fill_row_even
                day_num = (row_ptr - 7)
                ws_month.cell(row=row_ptr, column=1, value=f"Day {day_num:02d}").alignment = align_center
                ws_month.cell(row=row_ptr, column=5, value=False).alignment = align_center
                
                for col_i in range(1, 7):
                    cell = ws_month.cell(row=row_ptr, column=col_i)
                    cell.font = Font(name="Segoe UI", size=10, color="64748B")
                    cell.fill = r_fill
                    cell.border = grid_border
                row_ptr += 1

            # Conditional formatting: If column E is TRUE, highlight row nicely
            green_fill = PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid")
            green_font = Font(name="Segoe UI", size=10, color="34D399", bold=True)
            ws_month.conditional_formatting.add(
                "A8:F100",
                FormulaRule(formula=['$E8=TRUE'], stopIfTrue=True, fill=green_fill, font=green_font)
            )

            # Month column sizing
            ws_month.column_dimensions["A"].width = 14
            ws_month.column_dimensions["B"].width = 44
            ws_month.column_dimensions["C"].width = 16
            ws_month.column_dimensions["D"].width = 14
            ws_month.column_dimensions["E"].width = 24
            ws_month.column_dimensions["F"].width = 22
            ws_month.column_dimensions["G"].width = 12

        wb.save(self.filepath)
        logger.info(f"[ExcelTrackerEngine] Workbook successfully saved at '{self.filepath}'")
        return self.filepath


# Singleton excel engine
excel_tracker = ExcelTrackerEngine()
